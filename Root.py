# BIBLIOTECAS -----------------------------------------------------------------------------------

import PySimpleGUI as sg
import openpyxl as op



# TEMA DO PROGRAMA ------------------------------------------------------------------------------
sg.theme('Black')



# TELA DE LOGIN ---------------------------------------------------------------------------------

def sg_Login():
    layoutLogin = [
        [sg.Image(filename='./estacio.png')],
        [sg.Text('Usuário'),],
        [sg.Input(key='usuario')],
        [sg.Text('Senha')],
        [sg.Input(key='senha', password_char='*')],
        [sg.Text('', key='mensagem')],
        [sg.Button('Login', button_color='#5c2fd8', size=(10)), 
        sg.Button('Sair', button_color='#a0a0a2', size=(5)), sg.Button('Developer', button_color='#a0a0a0', size=(8))],
    ]
    
    janelaLogin = sg.Window('LOGIN', layout=layoutLogin, font='Roboto 15', size=(500,400))

    while True:
        event, values = janelaLogin.read()
        if event == 'Developer':
            janelaLogin.close()
            Developer()
        if event == 'Sair':
            sg.Exit()
            break
        elif event == sg.WIN_CLOSED:
            break
        elif event == 'Login':
#--------------------------------------------------------------
            usuario_correto = 'admin' #USUÁRIO DE ACESSO
            senha_correta = '1234' #SENHA DE ACESSO
#--------------------------------------------------------------            
            usuario = values['usuario']
            senha = values['senha']
            if senha == senha_correta and usuario == usuario_correto:
                janelaLogin['mensagem'].update('Login feito com sucesso.', text_color='#00FF00')
                janelaLogin.close()
                sg_Menu()
            else:
                janelaLogin['mensagem'].update('Senha ou Usuário incorreto', text_color='#ff0000')



# TELA DO MENU PRINCIPAL ---------------------------------------------------------------------------------

def sg_Menu():
    layoutMenu = [
        [sg.Text('CENTRAL FERRAMENTARIA', font='Roboto 30', text_color='#5c2fd8')],
        [sg.HSeparator()],
        [sg.Text('')],
        [sg.Button('CADASTRAR', size=(20),), sg.Button('CONSULTAR', size=(20))],
        [sg.Button('EDITAR', size=(20)),sg.Button('ESTOQUE', size=(20))],
        [sg.Text('')],
        [sg.Button('Sair', button_color='#a0a0a0', font='Roboto 15', size=(15))],
    ]
    
    janelaMenu = sg.Window('PROGRAMA - JEFFERSON', layout=layoutMenu, size=(500,300), font='Roboto 15', element_justification='center')
    
    while True:
        event, values = janelaMenu.read()
        if event == 'CADASTRAR':
            janelaMenu.close()
            sg_CadastroGeral()
        if event == 'CONSULTAR':
            janelaMenu.close()
            sg_Consultas()
        if event == 'EDITAR':
            janelaMenu.close()
            sg_Editar()
        if event == 'ESTOQUE':
            janelaMenu.close()
            sg_Estoque()
        if event == 'Sair':
            janelaMenu.close()
            sg.Exit()
            break
        elif event == sg.WIN_CLOSED:
            break



# TELA DOS CADASTROS ---------------------------------------------------------------------------------

def sg_CadastroGeral():
    layoutCadastroGeral = [
        [sg.Text('CADASTRAR', font='Roboto 30', text_color='#5c2fd8')],
        [sg.HSeparator()],
        [sg.Text()],
        [sg.Button('FERRAMENTAS', size=(20)),sg.Button('TECNICOS', size=(20))],
        [sg.Button('RESERVA', size=(20))],
        [sg.Text()],
        [sg.Button('Voltar', button_color='#a0a0a0')],
        ]
        
    janelaCadastroGeral = sg.Window('CADASTROS', layout=layoutCadastroGeral, size=(500,300), element_justification='center', font='Roboto 15')

    while True:
        event, values = janelaCadastroGeral.read()
        if event =='FERRAMENTAS':
            janelaCadastroGeral.close()
            sg_CadastroF()
        elif event =='TECNICOS':
            janelaCadastroGeral.close()
            sg_CadastroT()
        elif event =='RESERVA':
            janelaCadastroGeral.close()
            sg_CadastroR()
        elif event =='Voltar':
            janelaCadastroGeral.close()
            sg_Menu()
        elif event == sg.WIN_CLOSED:
            break



# TELA DO CADASTRO DE FERRAMENTAS --------------------------------------------------------------

def sg_CadastroF():   
    layoutCadF = [ 
                [sg.Text('Ferramenta'), sg.InputText('', key='FERRAMENTA')],
                [sg.Text('Modelo'), sg.InputText('', key='MODELO')],
                [sg.Text('Fabricante'), sg.InputText('', key='FABRICANTE')],
                [sg.Text('Material'), sg.InputText('', key='MATERIAL')],
                [sg.Text('Quantidade'), sg.InputText('', key='QUANTIDADE')],
                [sg.Text('Voltagem'), sg.InputText('', key='VOLTAGEM')],
                [sg.Text('Tipo'), sg.InputText('', key='TIPO')],
                [sg.Submit('Cadastrar', size=(10)), sg.Button('Voltar', button_color='#a0a0a0', size=(10))] 
    ]

    janelaCadastroF = sg.Window('CADASTRO DE FERRAMENTAS', layout=layoutCadF, font='Roboto 15', element_justification='center')
    
    while True:
        event, values = janelaCadastroF.read()
        if event == 'Cadastrar':
            sg.Popup('Ferramenta Cadastrada', font='Roboto 20', 
            background_color='#5c2fd8', 
            auto_close=True, 
            auto_close_duration=10,
            no_titlebar=True)
            lista = list(values.values())
            Add_Info('Ferramentas.xlsx', lista)
            janelaCadastroF.close()
            sg_CadastroF()
        if event == 'Voltar':
            janelaCadastroF.close()
            sg_CadastroGeral()
        if event == sg.WIN_CLOSED:
            sg.Exit()
            break



# TELA DO CADASTRO DE TÉCNICOS -----------------------------------------------------------------

def sg_CadastroT():
    layoutCadT = [ 
            [sg.Text('CPF'), sg.InputText(key='CPF')],
            [sg.Text('Nome'), sg.InputText(key='Nome')],
            [sg.Text('Sobrenome'), sg.InputText(key='Sobrenome')],
            [sg.Text('Telefone'), sg.InputText(key='Telefone')],
            [sg.Text('Turno: '), sg.Radio('Manhã', 'RADIO', key='MANHA'), sg.Radio('Noite', 'RADIO', key='Noite')],
            [sg.Text('Equipe: '), 
            sg.Radio('Team 1', 'RADIO', key='TEAM1'), 
            sg.Radio('Team 2', 'RADIO', key='TEAM2'),
            sg.Radio('Team 3', 'RADIO', key='TEAM3'),
            sg.Radio('Team 4', 'RADIO', key='TEAM4')],
            [sg.Submit('Cadastrar', size=(10)), sg.Button('Voltar', button_color='#a0a0a0', size=(10))] 
            ]

    janelaCadastroT = sg.Window('CADASTRO DE TÉCNICOS', layout=layoutCadT, font='Roboto 15', element_justification='center')
    
    while True:
        event, values = janelaCadastroT.read()
        if event == 'Cadastrar':
            sg.Popup('Técnico Cadastrado', font='Roboto 20', 
            background_color='#5c2fd8', 
            auto_close=True, 
            auto_close_duration=10,
            no_titlebar=True)
            lista = list(values.values())
            Add_Info('Tecnicos.xlsx', lista)
            janelaCadastroT.close()
            sg_CadastroT()
        if event == 'Voltar':
            janelaCadastroT.close()
            sg_CadastroGeral()
        if event == sg.WIN_CLOSED:
            sg.Exit()
            break



# TELA DO CADASTRO DE RESERVAS ---------------------------------------------------------------------------------------------------------------------------------------------------------

def sg_CadastroR():
    layoutCadR = [ 
            [sg.Text('Tecnico'), sg.InputText(key='TECNICO')],
            [sg.Text('Ferramenta'), sg.InputText(key='FERRAMENTA')],
            [sg.Input(key='DT-RESERVA', size=(20,1)) , sg.CalendarButton('Data da Reserva', close_when_date_chosen=True, target='DT-RESERVA',location=(800,400),no_titlebar=False)],
            [sg.Input(key='DT-DEVOLUCAO', size=(20,1)), sg.CalendarButton('Data da Devolução', close_when_date_chosen=True, target='DT-DEVOLUCAO',location=(800,400),no_titlebar=False)],
            [sg.Submit('Cadastrar', size=(10)), sg.Button('Voltar', button_color='#a0a0a0', size=(10))] 
            ]

    janelaCadastroR = sg.Window('CADASTRO DE RESERVAS', layout=layoutCadR, font='Roboto 15', element_justification='center')
    
    while True:
        event, values = janelaCadastroR.read()
        if event == 'Cadastrar':
            sg.Popup('Reserva Cadastrada', font='Roboto 20', 
            background_color='#5c2fd8', 
            auto_close=True, 
            auto_close_duration=10,
            no_titlebar=True)
            listaTemp = list(values.values()) #dados temporarios
            lista=[]
            for item in listaTemp: # verificar espaços 
                if item not in ['', None]:
                    lista.append(item)
            Add_Info('Reservas.xlsx', lista)
            janelaCadastroR.close()
            sg_CadastroR()
        if event == 'Voltar':
            janelaCadastroR.close()
            sg_CadastroGeral()
        if event == sg.WIN_CLOSED:
            sg.Exit()
            break



# TELA DE CONSULTAS -------------------------------------------------------------------------------------------------------------------------------

def sg_Consultas():
    layoutCadastroGeral = [
        [sg.Text('CONSULTAR', font='Roboto 30', text_color='#5c2fd8')],
        [sg.HSeparator()],
        [sg.Text()],
        [sg.Button('FERRAMENTAS', size=(20)), sg.Button('TECNICOS', size=(20))],
        [sg.Button('RESERVAS', size=(20))],
        [sg.Text()],
        [sg.Button('Voltar', button_color='#a0a0a0', size=(20))],
        ]
        
    janelaConsultas = sg.Window('CONSULTAS', layout=layoutCadastroGeral, size=(500,300), element_justification='center', font='Roboto 15')

    while True:
        event, values = janelaConsultas.read()
        if event =='FERRAMENTAS':
            janelaConsultas.close()
            sg_ConsultaF()
        elif event =='TECNICOS':
            janelaConsultas.close()
            sg_ConsultaT()
        elif event =='RESERVAS':
            janelaConsultas.close()
            sg_ConsultaR()
        elif event =='Voltar':
            janelaConsultas.close()
            sg_Menu()
        elif event == sg.WIN_CLOSED:
            break



# TELA DE CONSULTA DE FERRAMENTAS ---------------------------------------------------------------------------

def sg_ConsultaF(lista=False, values='', colunas=[1,2,3]):
    if lista == False:
        listaF=[]
    else:
        listaF = CarregarTabela('Ferramentas.xlsx', filtro=True, values = values, colunas = colunas)
    
    listaF_head = ['Ferramentas','Modelo','Fabricante','Material','Quantidade','Voltagem','Tipo']
    layoutConsultaF_topo = [ 
            [sg.Text('Ferramenta'), sg.InputText(key='-CS-Ferramenta-')],
            [sg.Text('Modelo'), sg.InputText(key='-CS-Modelo-')],
            [sg.Text('Fabricante'), sg.InputText(key='-CS-Fabricante-')],
            [sg.Submit('Consultar', size=(10)), sg.Button('Voltar', button_color='#a0a0a0', size=(10))],
            ]
    
    layoutConsultaF_baixo = [
        [sg.Table(values=listaF, headings=listaF_head ,max_col_width=35,
        auto_size_columns=True,
        justification='center',
        num_rows=10,
        key='-TABELA-',
        row_height=35)],
    ]

    layoutConsultaF = [
        [sg.Column(layoutConsultaF_topo)],
        [sg.HSeparator()],
        [sg.Column(layoutConsultaF_baixo)],

    ]

    janelaConsultaF = sg.Window('CONSULTA DE FERRAMENTAS', layout=layoutConsultaF, font='Roboto 15', element_justification='center')

    while True:
        event, values = janelaConsultaF.read()
        if event == 'Consultar':
            sg.popup('Cosulta Realizada',
            font='Roboto 20', 
            background_color='#5c2fd8', 
            auto_close=True, 
            auto_close_duration=10,
            no_titlebar=True)
            lista = list(values.values())
            janelaConsultaF.close()
            sg_ConsultaF(lista=True, values=lista)
        if event == 'Voltar':
            janelaConsultaF.close()
            sg_Consultas()
            break
        elif event == sg.WIN_CLOSED:
            break



# TELA DE CONSULTA DE TÉCNICOS -------------------------------------------------------------------------

def sg_ConsultaT(lista=False, values='', colunas=[1,2,3]):
    if lista == False:
        listaT=[]
    else:
        listaT = CarregarTabela('Tecnicos.xlsx', filtro=True, values = values, colunas = colunas)
    
    listaT_head = ['CPF','Nome','Sobrenome','Telefone','Turno']
    layoutConsultaT_topo = [ 
            [sg.Text('CPF'), sg.InputText(key='-CPF-')],
            [sg.Text('Nome'), sg.InputText(key='-NOME-')],
            [sg.Text('Sobrenome'), sg.InputText(key='-SOBRENOME-')],
            [sg.Submit('Consultar', size=(10)), sg.Button('Voltar', button_color='#a0a0a0', size=(10))],
            ]
    
    layoutConsultaT_baixo = [
        [sg.Table(values=listaT, headings=listaT_head ,max_col_width=35,
        auto_size_columns=True,
        justification='center',
        num_rows=10,
        key='-TABELA-',
        row_height=35)],
    ]

    layoutConsultaT = [
        [sg.Column(layoutConsultaT_topo)],
        [sg.HSeparator()],
        [sg.Column(layoutConsultaT_baixo)],
    ]

    janelaConsultaT = sg.Window('CONSULTA DE TÉCNICOS', layout=layoutConsultaT, font='Roboto 15', element_justification='center')

    while True:
        event, values = janelaConsultaT.read()
        if event == 'Consultar':
            sg.popup('Consulta Realizada',
            font='Roboto 20', 
            background_color='#5c2fd8', 
            auto_close=True, 
            auto_close_duration=10,
            no_titlebar=True)
            lista = list(values.values())
            janelaConsultaT.close()
            sg_ConsultaT(lista=True, values=lista)
        if event == 'Voltar':
            janelaConsultaT.close()
            sg_Consultas()
            break
        elif event == sg.WIN_CLOSED:
            break



# TELA DE CONSULTA DE RESERVAS -------------------------------------------------------------------------

def sg_ConsultaR(lista=False, values='', colunas=[1,2,3]):
    if lista == False:
        listaR=[]
    else:
        listaR = CarregarTabela('Reservas.xlsx', filtro=True, values = values, colunas = colunas)
    
    listaR_head = ['Técnico','Ferramenta','Reserva','Devolução']
    layoutConsultaR_topo = [ 
            [sg.Text('Técnico'), sg.InputText(key='-Tecnico-')],
            [sg.Text('Ferramenta'), sg.InputText(key='-Ferramenta-')],
            [sg.Input(key='DT-RESERVA', size=(20,1)) , sg.CalendarButton('Data da Reserva', close_when_date_chosen=True, target='DT-RESERVA',location=(800,400),no_titlebar=False)],
            [sg.Submit('Consultar', size=(10)), sg.Button('Voltar', button_color='#a0a0a0', size=(10))],
            ]
    
    layoutConsultaR_baixo = [
        [sg.Table(values=listaR, headings=listaR_head ,max_col_width=35,
        auto_size_columns=True,
        justification='center',
        num_rows=10,
        key='-TABELA-',
        row_height=35)],
    ]

    layoutConsultaR = [
        [sg.Column(layoutConsultaR_topo)],
        [sg.HSeparator()],
        [sg.Column(layoutConsultaR_baixo)],
    ]

    janelaConsultaR = sg.Window('CONSULTA DE RESERVAS', layout=layoutConsultaR, font='Roboto 15', element_justification='center')

    while True:
        event, values = janelaConsultaR.read()
        if event == 'Consultar':
            sg.popup('Consulta Realizada',
            font='Roboto 20', 
            background_color='#5c2fd8', 
            auto_close=True, 
            auto_close_duration=10,
            no_titlebar=True)
            lista = list(values.values())
            janelaConsultaR.close()
            sg_ConsultaR(lista=True, values=lista)
        if event == 'Voltar':
            janelaConsultaR.close()
            sg_Consultas()
            break
        elif event == sg.WIN_CLOSED:
            break



# TELA DE EDIÇÕES --------------------------------------------------------------------------------------

def sg_Editar():
    layoutCadastroGeral = [
        [sg.Text('EDITAR', font='Roboto 30', text_color='#5c2fd8')],
        [sg.HSeparator()],
        [sg.Text()],
        [sg.Button('FERRAMENTAS', size=(20)), sg.Button('TECNICOS', size=(20))],
        [sg.Button('RESERVAS', size=(20))],
        [sg.Text()],
        [sg.Button('Voltar', button_color='#a0a0a0')],
        ]
        
    janelaCadastroGeral = sg.Window('EDITAR', layout=layoutCadastroGeral, size=(500,300), element_justification='center', font='Roboto 15')

    while True:
        event, values = janelaCadastroGeral.read()
        if event =='FERRAMENTAS':
            janelaCadastroGeral.close()
            sg_Editar()
        elif event =='TECNICOS':
            janelaCadastroGeral.close()
            sg_Editar()
        elif event =='RESERVA':
            janelaCadastroGeral.close()
            sg_Editar()
        elif event =='Voltar':
            janelaCadastroGeral.close()
            sg_Menu()
        elif event == sg.WIN_CLOSED:
            break




# BANCO DE DADOS ------------------------------------------------------------

def AbrirBD(caminho):
    bd = op.load_workbook(caminho)
    return bd

def FecharBD(caminho, bd):
    bd.save(caminho)
    bd.close()


def Add_Info(caminho, lista):
    wb = AbrirBD(caminho)
    ws = wb['Sheet1']
    ultimaL = 1
    for lin in range(1,1000000): # verificação das linhas
        if ws.cell(row=lin, column=1).value == None:
            ultimaL = lin
            break
    for col in range(1, len(lista)+1): # Add valores na coluna
        ws.cell(row=ultimaL, column=col).value = lista[col-1]

    FecharBD(caminho, wb)



# FILTROS DE PESQUISA - Tive ajudar do @Samuel-Borba de outro grupo ------------------------------------------------------------------

def CarregarTabela(caminho, values='', filtro=False, colunas=''):
    wb = AbrirBD(caminho)
    ws = wb['Sheet1']
    tabela = []
    linha = []
    checkFiltro = True
    for lin in range(2,1000000):
        if ws.cell(row=lin, column=1).value == None: break
        linha = []
        checkFiltro = True
        if filtro == True:
            cont = 0
            for col in colunas:
                if values[cont] in ['', []]:
                    cont += 1
                    continue
                if str(ws.cell(row=lin, column=col).value).find(values[cont]) == -1:  # verificando coluna via filtro
                    checkFiltro = False
                    break
                cont += 1
        if checkFiltro == False : continue

        for col in range(1,1000):
            if ws.cell(row=lin, column=col).value == None: break
            linha.append(ws.cell(row=lin, column=col).value)
        tabela.append(linha)

    FecharBD(caminho, wb)
    return tabela



# TELA ESTOQUE -------------------------------------------------------------------------------------------------------

def sg_Estoque(lista=False, values='', colunas=[1,2,3,4,5,6,7]):
    if lista == False:
        listaE=[]
    else:
        listaE = CarregarTabela('Ferramentas.xlsx', filtro=False, values = values, colunas = colunas)
    
    tabela_head = ['Ferramentas','Modelo','Fabricante','Material','Quantidade','Voltagem','Tipo']
    layoutEstoque_top = [ 
            [sg.Text('ESTOQUE DE FERRAMENTAS'), sg.Submit('Listar', size=(10)), sg.Button('Voltar', button_color='#a0a0a0', size=(10))],
            ]
    
    layoutEstoque_bot = [
        [sg.Table(values=listaE, headings=tabela_head ,max_col_width=35,
        auto_size_columns=True,
        justification='center',
        num_rows=10,
        key='-TABELA-',
        row_height=35)],
    ]

    layoutEstoque = [
        [sg.Column(layoutEstoque_top)],
        [sg.HSeparator()],
        [sg.Column(layoutEstoque_bot)],

    ]

    janelaEstoque = sg.Window('ESTOQUE', layout=layoutEstoque, font='Roboto 15', element_justification='center')

    while True:
        event, values = janelaEstoque.read()
        if event == 'Listar':
            sg.popup('Listar Estoque de Ferramentas',
            font='Roboto 20', 
            background_color='#5c2fd8', 
            auto_close=True, 
            auto_close_duration=10,
            no_titlebar=True)
            lista = list(values.values())
            janelaEstoque.close()
            sg_Estoque(lista=True, values=values)
        if event == 'Voltar':
            janelaEstoque.close()
            sg_Menu()
            break
        elif event == sg.WIN_CLOSED:
            break



# TELA DEVELOPER ----------------------------------------------------------

def Developer():
    layoutDEV = [
        [sg.Image(filename='./dev.png')],
        [sg.HSeparator()],
        [sg.Text('Desenvolvedor: Jefferson Ponte Pessoa')],
        [sg.Text('Matricula: 202208291228')],
        [sg.Text('Curso: Desenvolvimento Full Stack')],
        [sg.Text('Turma: 22.3')],
        [sg.Push()],

        sg.Button('Sair', button_color='#5c2fd8', size=(10)),
        ],
    
    janelaDEV = sg.Window('Developer', layout=layoutDEV, font='Roboto 20', element_justification='center', size=(500,300))

    while True:
        event, values = janelaDEV.read()
        if event == 'Sair':
            janelaDEV.close()
            sg_Login()
        elif event == sg.WIN_CLOSED:
            break









