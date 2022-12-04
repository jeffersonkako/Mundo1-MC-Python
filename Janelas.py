import PySimpleGUI as sg
import pandas as pd
import openpyxl as op

# Tema do programa ------------------------------------------------------------------------------
sg.theme('Black')


# tela de login ---------------------------------------------------------------------------------
def sg_Login():
    layoutLogin = [
        [sg.Text('Usuário')],
        [sg.Input(key='usuario')],
        [sg.Text('Senha')],
        [sg.Input(key='senha', password_char='*')],
        [sg.Text('', key='mensagem')],
        [sg.Button('Login', button_color='#5c2fd8'), 
        sg.Button('Sair', button_color='#a0a0a0'), sg.Push(),
        sg.Button('Equipe', button_color='#a0a0a0')
        ],
    ]
    
    janelaLogin = sg.Window('LOGIN', layout=layoutLogin, font='Roboto 15')

    while True:
        event, values = janelaLogin.read()
        if event == 'Equipe':
            Team2_inf = DadosEquipe()
            sg.popup(Team2_inf,
            font='Roboto 20', 
            background_color='#5c2fd8', 
            auto_close=True, 
            auto_close_duration=10,
            no_titlebar=True)
            janelaLogin.close()
            sg_Login()
        if event == 'Sair':
            sg.Exit()
            break
        elif event == sg.WIN_CLOSED:
            break
        elif event == 'Login':
            usuario_correto = 'admin'
            senha_correta = '1234'
            usuario = values['usuario']
            senha = values['senha']
            if senha == senha_correta and usuario == usuario_correto:
                janelaLogin['mensagem'].update('Login feito com sucesso.', text_color='#00FF00')
                janelaLogin.close()
                sg_Menu()
            else:
                janelaLogin['mensagem'].update('Senha ou Usuário incorreto', text_color='#ff0000')


# TELA DO MENU ---------------------------------------------------------------------------------
def sg_Menu():
    layoutMenu = [
        [sg.Text('Menu Principal', font='Roboto 30', text_color='#5c2fd8')],
        [sg.HSeparator()],
        [sg.Text()],
        [sg.Button('CADASTRAR'), sg.Button('CONSULTAR')],
        [sg.Button('EDITAR')],
        [sg.Text()],
        [sg.Button('Sair', button_color='#a0a0a0', font='Roboto 15')],
    ]
    
    janelaMenu = sg.Window('PROGRAMA - TEAM 02', layout=layoutMenu, size=(400,300), element_justification='center', font='Roboto 15')
    
    while True:
        event, values = janelaMenu.read()
        if event == 'CADASTRAR':
            janelaMenu.close()
            sg_CadastroGeral()
        if event == 'CONSULTAR':
            janelaMenu.close()
            sg_Consultas()
        if event == 'EDITAR':
            janelaMenu.close()
            sg_Editar()
        if event == 'Sair':
            janelaMenu.close()
            sg.Exit()
            break
        elif event == sg.WIN_CLOSED:
            break


# TELA DO CADASTRO---------------------------------------------------------------------------------
def sg_CadastroGeral():
    layoutCadastroGeral = [
        [sg.Text('CADASTRAR', font='Roboto 30', text_color='#5c2fd8')],
        [sg.HSeparator()],
        [sg.Text()],
        [sg.Button('FERRAMENTAS')],
        [sg.Button('TECNICOS')],
        [sg.Button('RESERVA')],
        [sg.Text()],
        [sg.Button('Voltar', button_color='#a0a0a0')],
        ]
        
    janelaCadastroGeral = sg.Window('CADASTRO', layout=layoutCadastroGeral, size=(400,300), element_justification='center', font='Roboto 15')

    while True:
        event, values = janelaCadastroGeral.read()
        if event =='FERRAMENTAS':
            janelaCadastroGeral.close()
            sg_CadastroF()
        elif event =='TECNICOS':
            janelaCadastroGeral.close()
            sg_CadastroT()
        elif event =='RESERVA':
            janelaCadastroGeral.close()
            sg_CadastroR()
        elif event =='Voltar':
            janelaCadastroGeral.close()
            sg_Menu()
        elif event == sg.WIN_CLOSED:
            break


# TELA DO CADASTRO DE FERRAMENTAS --------------------------------------------------------------
def sg_CadastroF():   
    layoutCadF = [ 
                [sg.Text('Ferramenta'), sg.InputText('', key='FERRAMENTA')],
                [sg.Text('Modelo'), sg.InputText('', key='MODELO')],
                [sg.Text('Fabricante'), sg.InputText('', key='FABRICANTE')],
                [sg.Text('Peso'), sg.InputText('', key='PESO')],
                [sg.Text('Quantidade'), sg.InputText('', key='QUANTIDADE')],
                [sg.Text('Voltagem'), sg.InputText('', key='VOLTAGEM')],
                [sg.Text('Tipo'), sg.InputText('', key='TIPO')],
                [sg.Submit('Cadastrar'), sg.Button('Voltar', button_color='#a0a0a0')] 
    ]

    janelaCadastroF = sg.Window('CADASTRO DE FERRAMENTAS', layout=layoutCadF, font='Roboto 15')
    
    while True:
        event, values = janelaCadastroF.read()
        if event == 'Cadastrar':
            sg.Popup('Ferramenta Cadastrada', font='Roboto 20', 
            background_color='#007d4d', 
            auto_close=True, 
            auto_close_duration=10,
            no_titlebar=True)
            lista = list(values.values())
            Add_Info('Ferramentas.xlsx', lista)
            janelaCadastroF.close()
            sg_CadastroF()
        if event == 'Voltar':
            janelaCadastroF.close()
            sg_CadastroGeral()
        if event == sg.WIN_CLOSED:
            sg.exit()
            break



# TELA DO CADASTRO DE TÉCNICOS -----------------------------------------------------------------
def sg_CadastroT():
    layoutCadT = [ 
            [sg.Text('CPF'), sg.InputText(key='CPF')],
            [sg.Text('Nome'), sg.InputText(key='Nome')],
            [sg.Text('Sobrenome'), sg.InputText(key='Sobrenome')],
            [sg.Text('Telefone'), sg.InputText(key='Telefone')],
            [sg.Text('Turno: '), sg.Radio('Manhã', 'RADIO', key='MANHA'), sg.Radio('Noite', 'RADIO', key='Noite')],
            [sg.Text('Equipe: '), 
            sg.Radio('Team 1', 'RADIO', key='TEAM1'), 
            sg.Radio('Team 2', 'RADIO', key='TEAM2'),
            sg.Radio('Team 3', 'RADIO', key='TEAM3'),
            sg.Radio('Team 4', 'RADIO', key='TEAM4')],
            [sg.Submit('Cadastrar'), sg.Button('Voltar', button_color='#a0a0a0')] 
            ]

    janelaCadastroT = sg.Window('CADASTRO DE TÉCNICOS', layout=layoutCadT, font='Roboto 15')
    
    while True:
        event, values = janelaCadastroT.read()
        if event == 'Cadastrar':
            sg.Popup('Técnico Cadastrado', font='Roboto 20', 
            background_color='#007d4d', 
            auto_close=True, 
            auto_close_duration=10,
            no_titlebar=True)
            lista = list(values.values())
            Add_Info('Tecnicos.xlsx', lista)
            janelaCadastroT.close()
            sg_CadastroF()
        if event == 'Voltar':
            janelaCadastroT.close()
            sg_CadastroGeral()
        if event == sg.WIN_CLOSED:
            sg.exit()
            break


# TELA DO CADASTRO DE RESERVAS ---------------------------------------------------------------------------------------------------------------------------------------------------------
def sg_CadastroR():
    layoutCadR = [ 
            [sg.Text('Tecnico'), sg.InputText(key='TECNICO')],
            [sg.Text('Ferramenta'), sg.InputText(key='FERRAMENTA')],
            [sg.Input(key='DT-RESERVA', size=(20,1)) , sg.CalendarButton('Data da Reserva', close_when_date_chosen=True, target='DT-RESERVA',location=(800,400),no_titlebar=False)],
            [sg.Input(key='DT-DEVOLUCAO', size=(20,1)), sg.CalendarButton('Data da Devolução', close_when_date_chosen=True, target='DT-DEVOLUCAO',location=(800,400),no_titlebar=False)],
            [sg.Submit('Cadastrar'), sg.Button('Voltar', button_color='#a0a0a0')] 
            ]

    janelaCadastroR = sg.Window('CADASTRO DE RESERVAS', layout=layoutCadR, font='Roboto 15')
    
    while True:
        event, values = janelaCadastroR.read()
        if event == 'Cadastrar':
            sg.Popup('Reserva Cadastrada', font='Roboto 20', 
            background_color='#007d4d', 
            auto_close=True, 
            auto_close_duration=10,
            no_titlebar=True)
            lista = list(values.values())
            Add_Info('Reservas.xlsx', lista)
            janelaCadastroR.close()
            sg_CadastroF()
        if event == 'Voltar':
            janelaCadastroR.close()
            sg_CadastroGeral()
        if event == sg.WIN_CLOSED:
            sg.exit()
            break

# TELA DE CONSULTAS -------------------------------------------------------------------------------------------------------------------------------
def sg_Consultas():
    layoutCadastroGeral = [
        [sg.Text('CONSULTAR', font='Roboto 30', text_color='#5c2fd8')],
        [sg.HSeparator()],
        [sg.Text()],
        [sg.Button('FERRAMENTAS')],
        [sg.Button('TECNICOS')],
        [sg.Button('RESERVAS')],
        [sg.Text()],
        [sg.Button('Voltar', button_color='#a0a0a0')],
        ]
        
    janelaConsultas = sg.Window('CONSULTAS', layout=layoutCadastroGeral, size=(400,300), element_justification='center', font='Roboto 15')

    while True:
        event, values = janelaConsultas.read()
        if event =='FERRAMENTAS':
            janelaConsultas.close()
            sg_ConsultaF()
        elif event =='TECNICOS':
            janelaConsultas.close()
            sg_Consultas()
        elif event =='RESERVAS':
            janelaConsultas.close()
            sg_Consultas()
        elif event =='Voltar':
            janelaConsultas.close()
            sg_Menu()
        elif event == sg.WIN_CLOSED:
            break

# TELA DE REALIZAR CONSULTAS ---------------------------------------------------------------------------
listaF = []
listaF_head = ['Ferramentas','Modelo','Fabricante','Peso','Quantidade','Voltagem','Tipo']
def sg_ConsultaF():
    layoutConsultaF_ldE = [ 
            [sg.Text('Ferramenta'), sg.InputText(key='-CS-Ferramenta-')],
            [sg.Text('Modelo'), sg.InputText(key='-CS-Modelo-')],
            [sg.Text('Fabricante'), sg.InputText(key='-CS-Fabricante-')],
            [sg.Text('Peso'), sg.InputText(key='-CS-Peso-')],
            [sg.Button('Consultar'), sg.Button('Voltar', button_color='#a0a0a0')],
            ]
    
    layoutConsultaF_ldD = [
        [sg.Table(values=listaF, headings=listaF_head ,max_col_width=35,
        auto_size_columns=True,
        justification='center',
        num_rows=10,
        key='-TABELA-',
        row_height=35)],
    ]

    layoutConsultaF = [
        [sg.Column(layoutConsultaF_ldE)],
        [sg.HSeparator()],
        [sg.Column(layoutConsultaF_ldD)],

    ]

    janelaConsultaF = sg.Window('CADASTRO DE FERRAMENTAS', layout=layoutConsultaF, font='Roboto 15')

    while True:
        event, values = janelaConsultaF.read()
        if event == 'Consultar':
            sg.popup('Cosulta Realizada',
            font='Roboto 20', 
            background_color='#007d4d', 
            auto_close=True, 
            auto_close_duration=10,
            no_titlebar=True)
            janelaConsultaF.close()
            sg_ConsultaF()
        if event == 'Voltar':
            janelaConsultaF.close()
            sg_Consultas()
            break
        elif event == sg.WIN_CLOSED:
            break

# TELA DE EDIÇÕES --------------------------------------------------------------------------------------
def sg_Editar():
    layoutCadastroGeral = [
        [sg.Text('EDITAR', font='Roboto 30', text_color='#5c2fd8')],
        [sg.HSeparator()],
        [sg.Text()],
        [sg.Button('FERRAMENTAS')],
        [sg.Button('TECNICOS')],
        [sg.Button('RESERVAS')],
        [sg.Text()],
        [sg.Button('Voltar', button_color='#a0a0a0')],
        ]
        
    janelaCadastroGeral = sg.Window('EDITAR', layout=layoutCadastroGeral, size=(400,300), element_justification='center', font='Roboto 15')

    while True:
        event, values = janelaCadastroGeral.read()
        if event =='FERRAMENTAS':
            janelaCadastroGeral.close()
            sg_Editar()
        elif event =='TECNICOS':
            janelaCadastroGeral.close()
            sg_Editar()
        elif event =='RESERVA':
            janelaCadastroGeral.close()
            sg_Editar()
        elif event =='Voltar':
            janelaCadastroGeral.close()
            sg_Menu()
        elif event == sg.WIN_CLOSED:
            break


# DADOS INPUTS ---------------------------------------------------------------------------------------------
def DadosReserva(values):
    info = 'RESERVA CADASTRADA\n'
    nome = '\nNome: '+ values['R-TECNICO']
    info += nome
    ferramentaR = '\nFerramenta: '+ values['R-FERRAMENTA']
    info += ferramentaR
    dataR = '\nReserva: ' + values['DT-RESERVA']
    info += dataR
    dataD = '\nDevolução: ' + values['DT-DEVOLUCAO']
    info += dataD
    return info


def DadosFerramentas(values):
    info = 'FERRAMENTA CADASTRADA\n'
    Ferramenta = '\nFerramenta: '+ values['C-FERRAMENTAS']
    info += Ferramenta
    Fabricante = '\nFabricante: '+ values['C-FABRICANTE']
    info += Fabricante
    Peso = '\nPeso: ' + values['C-PESO']
    info += Peso
    Quantidade = '\nQuantidade: ' + values['C-QUANTIDADE']
    info += Quantidade
    Voltagem = '\nVoltagem: ' + values['C-VOLTAGENS']
    info += Voltagem
    Tipo = '\nTipo: ' + values['C-TIPO']
    info += Tipo
    return info


def DadosTecnico(values):
    info = 'Técnico Cadastrado\n'
    CPF = '\nCPF: '+ values['T-CPF']
    info += CPF
    Nome = '\nNome: '+ values['T-NOME']
    info += Nome
    Sobrenome = '\nSobrenome: ' + values['T-SOBRENOME']
    info += Sobrenome
    Telefone = '\nTelefone: ' + values['T-TELEFONE']
    info += Telefone
    return info

def DadosEquipe():
    info = 'Estacio - TEAM 2\n'
    Aluno1 = '\nJefferson'
    info += Aluno1
    Aluno2 = '\nAmandio'
    info += Aluno2
    Aluno3 = '\nAriel'
    info += Aluno3
    Aluno4 = '\nRaiza'
    info += Aluno4
    return info


# acessar banco de dados ------------------------------------------------------------

def AbrirBD(caminho):
    bd = op.load_workbook(caminho)
    return bd

def FecharBD(caminho, bd):
    bd.save(caminho)
    bd.close()


def Add_Info(caminho, lista):
    wb = AbrirBD(caminho)
    ws = wb['Sheet1']
    ultimaL = 1
    for lin in range(1,1000000): # verificação das linhas
        if ws.cell(row=lin, column=1).value == None:
            ultimaL = lin
            break
    for col in range(1, len(lista)+1): # Add valores na coluna
        ws.cell(row=ultimaL, column=col).value = lista[col-1]

    FecharBD(caminho, wb)







